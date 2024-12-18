import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries
import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess
import win32com.client
import pdfplumber
import re


#Processar arquivos
def processar_arquivo_csv(excel_path, extrato_csv, entry_planilha, tabela_entrada, tabela_saida):


    extrato = pd.read_csv(extrato_csv, delimiter=';', skiprows=3)

    wb = load_workbook(excel_path)


    # Remover espaços em branco das colunas, caso existam
    extrato.columns = extrato.columns.str.strip()

    # Substituir ponto (separador de milhar) e vírgula (separador decimal)
    extrato['Valor'] = extrato['Valor'].str.replace('.', '', regex=False)  # Remove pontos de milhares
    extrato['Valor'] = extrato['Valor'].str.replace(',', '.', regex=False)  # Substitui vírgula por ponto para separação decimal

    # Converter a coluna 'Valor' para float
    extrato['Valor'] = extrato['Valor'].astype(float)

    # Filtrar apenas as colunas de interesse: 'Descrição' (nome da pessoa) e 'Valor'
    dados_entradas = extrato[extrato['Valor'] >= 0][['Descrição', 'Valor']]
    dados_saidas = extrato[extrato['Valor'] < 0][['Descrição', 'Valor']]


    nome_planilha = entry_planilha.get()

    # teste de erro para planilha
    if nome_planilha == '':
        messagebox.showerror("Erro","Por favor, coloque o nome da planilha a ser inserida")
        return
    

    planilha_encontrada = False
    for planilha_erro in wb.worksheets: 
        if planilha_erro.title == nome_planilha:
            ws = wb[nome_planilha]
            planilha_encontrada = True
            break


    if not planilha_encontrada:
        messagebox.showerror("Erro",f"Planilha {nome_planilha} não encontrada")
        return

    nome_tabela_entrada = tabela_entrada.get()
    nome_tabela_saida = tabela_saida.get()

    # Teste de erro para as tabelas
    if nome_tabela_entrada == '':
        messagebox.showerror("Erro","Por favor, coloque o nome da tabela de entradas")
        return
    if nome_tabela_saida == '':
        messagebox.showerror("Erro","Por favor, coloque o nome da tabela de saidas")
        return

    # Adicionar entradas e saídas nas tabelas existentes
    erro_entrada = adicionar_dados_na_tabela(dados_entradas, ws, nome_tabela_entrada)  # Altere para o nome da tabela de entradas
    if not erro_entrada:
        return
    
    erro_saida = adicionar_dados_na_tabela(dados_saidas, ws, nome_tabela_saida)        # Altere para o nome da tabela de saídas

    if not erro_saida:
        return

    # Salvar as modificações no arquivo Excel
    wb.save(excel_path)

    messagebox.showinfo("Sucesso", "Dados adicionados com sucesso às tabelas existentes!")

    os.startfile(excel_path)
    return




# Função para adicionar os dados a uma tabela existente em uma aba específica
def adicionar_dados_na_tabela(dados, worksheet, tabela_nome):
    """
    Adiciona os dados na tabela estruturada existente.
    `dados`: DataFrame com os dados para adicionar.
    `worksheet`: Worksheet do openpyxl onde a tabela está localizada.
    `tabela_nome`: O nome da tabela estruturada dentro da aba.
    """
    # Encontrar a tabela dentro da planilha
    
    tabela_encontrada = False
    tabela = None
    for table in worksheet._tables:  # Acessa todas as tabelas na planilha
        if table == tabela_nome:  # Verifica se 'table' é uma tabela e compara o nome
            tabela = table  # Encontra a tabela
            tabela_encontrada = True
            break
    
    if not tabela_encontrada:
        messagebox.showerror("Erro",f"Tabela {tabela_nome} não encontrada")
        return False


    # Encontrar a primeira linha vazia após a tabela existente
    ref = worksheet._tables[tabela].ref  # Exemplo: 'A1:B5'
    start_col, start_row, end_col, end_row = range_boundaries(ref)
    first_empty_row = start_row + 1

    # Adicionar os dados a partir da primeira linha vazia
    for i, row in enumerate(dados.itertuples(index=False), start=first_empty_row):
        worksheet.cell(row=i, column=start_col, value=row.Descrição)  # Acessa a coluna 'Descrição'
        worksheet.cell(row=i, column=start_col + 1, value=row.Valor)  # Acessa a coluna 'Valor'
    return True
        


def fechar_arquivo(caminho_xlsx, caminho_arquivo, entry_planilha, tabela_entrada, tabela_saida):
    
    try:
        excel = win32com.client.Dispatch("Excel.Application")

    except Exception as e:
        print(f"Erro ao inicializar o excel: {e}")
        return
    
    try:
        arquivo_aberto = False

        # Verificar se o Excel está aberto
        for workbook in excel.Workbooks:
            if os.path.abspath(workbook.FullName) == os.path.abspath(caminho_xlsx):
                print(f"O arquivo {workbook.FullName} foi fechado")
                workbook.Close(SaveChanges=False)
                arquivo_aberto = True
                break

        if caminho_arquivo and caminho_xlsx:
            if not arquivo_aberto:
                print("Arquivo não esta aberto")
            
            
            extensao = os.path.splitext(caminho_arquivo)[1]

            if extensao in '.csv':
                processar_arquivo_csv(caminho_xlsx, caminho_arquivo, entry_planilha, tabela_entrada, tabela_saida)
                return
            elif extensao in '.pdf':
                processar_arquivo_pdf(caminho_xlsx, caminho_arquivo, entry_planilha, tabela_entrada, tabela_saida)
                return
            else:
                print("extençâo não encontrada")
                return
    
    except Exception as e:
        messagebox.showerror("Erro",f"Erro ao Processar os arquivo: {e}")


def processar_arquivo_pdf(excel_path, extrato_pdf, entry_planilha, tabela_entrada, tabela_saida):
    dados_entradas, dados_saidas = ler_pdf(extrato_pdf) # Lógica para ler o pdf

    wb = load_workbook(excel_path)


    nome_planilha = entry_planilha.get()

    # teste de erro para planilha
    if nome_planilha == '':
        messagebox.showerror("Erro","Por favor, coloque o nome da planilha a ser inserida")
        return
    

    planilha_encontrada = False
    for planilha_erro in wb.worksheets: 
        if planilha_erro.title == nome_planilha:
            ws = wb[nome_planilha]
            planilha_encontrada = True
            break


    if not planilha_encontrada:
        messagebox.showerror("Erro",f"Planilha {nome_planilha} não encontrada")
        return

    nome_tabela_entrada = tabela_entrada.get()
    nome_tabela_saida = tabela_saida.get()

    # Teste de erro para as tabelas
    if nome_tabela_entrada == '':
        messagebox.showerror("Erro","Por favor, coloque o nome da tabela de entradas")
        return
    if nome_tabela_saida == '':
        messagebox.showerror("Erro","Por favor, coloque o nome da tabela de saidas")
        return

    # Adicionar entradas e saídas nas tabelas existentes
    erro_entrada = adicionar_dados_na_tabela(dados_entradas, ws, nome_tabela_entrada)  # Altere para o nome da tabela de entradas
    if not erro_entrada:
        return
    
    erro_saida = adicionar_dados_na_tabela(dados_saidas, ws, nome_tabela_saida)        # Altere para o nome da tabela de saídas

    if not erro_saida:
        return

    # Salvar as modificações no arquivo Excel
    wb.save(excel_path)

    messagebox.showinfo("Sucesso", "Dados adicionados com sucesso às tabelas existentes!")

    os.startfile(excel_path)
    return



def limpar_valor(valor):
    try:
        valor_limpo = re.sub(r'[^\d.,-]', '', valor).replace(',', '.').strip()
        if valor_limpo:
            return valor_limpo
        else:
            return '0'
    except Exception as e:
        print(f"Erro ao limpar valor: {valor} - {e}")
        return '0'

def ler_pdf(arquivo):
    try:
        with pdfplumber.open(arquivo) as pdf:
            entradas_paginas = []
            saidas_paginas = []

            for numero_pagina, pagina in enumerate(pdf.pages):
                tabela = pagina.extract_table()
                if tabela:
                    df = pd.DataFrame(tabela[1:], columns=tabela[0])

                    # Limpeza de valores
                    df['Valor'] = df['Valor'].apply(limpar_valor)
                    df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')

                    # Renomeia a coluna "Descrição das Movimentações" para "Descrição" se existir
                    if "Descrição das Movimentações" in df.columns:
                        df.rename(columns={"Descrição das Movimentações": "Descrição"}, inplace=True)

                    # Filtra valores positivos e negativos
                    entrada_pagina = df[df['Valor'] >= 0][["Descrição", 'Valor']]
                    saida_pagina = df[df['Valor'] < 0][["Descrição", 'Valor']]

                    entradas_paginas.append(entrada_pagina)
                    saidas_paginas.append(saida_pagina)

            entrada_consolidada = pd.concat(entradas_paginas, ignore_index=True) if entradas_paginas else pd.DataFrame()
            saida_consolidada = pd.concat(saidas_paginas, ignore_index=True) if saidas_paginas else pd.DataFrame()
            
            return entrada_consolidada, saida_consolidada

    except Exception as e:
        print(e)
        messagebox.showerror("Erro", f"Erro ao ler o PDF: {e}")